from __future__ import annotations

import json
import math
import os
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"
ARCHIVE_ROOT = Path(os.environ.get("RHEOLOGY_ARCHIVE_ROOT", ROOT / "outputs"))
RUN_ID = os.environ.get("RHEOLOGY_TABLE_RUN_ID") or datetime.now().strftime("%Y%m%d_%H%M%S")
OUT_DIR = OUTPUTS / f"publication_metric_table_{RUN_ID}"


MODEL_LABELS = {
    "GPR_Matern_ARD": "GPR-Matern-ARD",
    "KernelRidge_RBF": "Kernel ridge-RBF",
    "SVR_RBF": "SVR-RBF",
    "RandomForest": "Random forest",
    "GradientBoosting": "Gradient boosting",
}

MODEL_ORDER = [
    "GPR-Matern-ARD",
    "Kernel ridge-RBF",
    "SVR-RBF",
    "Ridge",
    "XGBoost",
    "ExtraTrees",
    "Random forest",
    "Gradient boosting",
]

PARAMETER_INFO = {
    "full_viscosity_curve|log_viscosity": {
        "section": "Viscosity parameters",
        "parameter": "Full viscosity curve",
        "symbol_unit": "log10 η(γ̇), cP = mPa·s; γ̇ in s⁻¹",
        "input_set": "formulation + log10 γ̇",
        "order": 101,
    },
    "viscosity_scalar|log10_eta_1": {
        "section": "Viscosity parameters",
        "parameter": "Viscosity at 1 s⁻¹",
        "symbol_unit": "log10 η1, cP = mPa·s",
        "input_set": "formulation only",
        "order": 102,
    },
    "viscosity_scalar|log10_eta_50": {
        "section": "Viscosity parameters",
        "parameter": "Viscosity at 50 s⁻¹",
        "symbol_unit": "log10 η50, cP = mPa·s",
        "input_set": "formulation only",
        "order": 103,
    },
    "viscosity_scalar|log10_eta_100": {
        "section": "Viscosity parameters",
        "parameter": "Viscosity at 100 s⁻¹",
        "symbol_unit": "log10 η100, cP = mPa·s",
        "input_set": "formulation only",
        "order": 104,
    },
    "viscosity_scalar|viscosity_shear_thinning_slope_1to100": {
        "section": "Viscosity parameters",
        "parameter": "Shear-thinning slope",
        "symbol_unit": "slope of log10 η vs log10 γ̇ from 1-100 s⁻¹",
        "input_set": "formulation only",
        "order": 105,
    },
    "saos_scalar|log10_Gp_0.1Hz": {
        "section": "SAOS parameters",
        "parameter": "Storage modulus at 0.1 Hz",
        "symbol_unit": "log10 G′0.1Hz, Pa",
        "input_set": "formulation only",
        "order": 1,
    },
    "saos_scalar|log10_Gpp_0.1Hz": {
        "section": "SAOS parameters",
        "parameter": "Loss modulus at 0.1 Hz",
        "symbol_unit": "log10 G″0.1Hz, Pa",
        "input_set": "formulation only",
        "order": 2,
    },
    "saos_scalar|tan_delta_0.1Hz": {
        "section": "SAOS parameters",
        "parameter": "Loss tangent at 0.1 Hz",
        "symbol_unit": "tan δ0.1Hz, dimensionless",
        "input_set": "formulation only",
        "order": 3,
    },
    "saos_scalar|log10_Gp_1Hz": {
        "section": "SAOS parameters",
        "parameter": "Storage modulus at 1 Hz",
        "symbol_unit": "log10 G′1Hz, Pa",
        "input_set": "formulation only",
        "order": 4,
    },
    "saos_scalar|log10_Gpp_1Hz": {
        "section": "SAOS parameters",
        "parameter": "Loss modulus at 1 Hz",
        "symbol_unit": "log10 G″1Hz, Pa",
        "input_set": "formulation only",
        "order": 5,
    },
    "saos_scalar|tan_delta_1Hz": {
        "section": "SAOS parameters",
        "parameter": "Loss tangent at 1 Hz",
        "symbol_unit": "tan δ1Hz, dimensionless",
        "input_set": "formulation only",
        "order": 6,
    },
    "saos_scalar|log10_Gp_6.31Hz": {
        "section": "SAOS parameters",
        "parameter": "Storage modulus at 6.31 Hz",
        "symbol_unit": "log10 G′6.31Hz, Pa",
        "input_set": "formulation only",
        "order": 7,
    },
    "saos_scalar|log10_Gpp_6.31Hz": {
        "section": "SAOS parameters",
        "parameter": "Loss modulus at 6.31 Hz",
        "symbol_unit": "log10 G″6.31Hz, Pa",
        "input_set": "formulation only",
        "order": 8,
    },
    "saos_scalar|tan_delta_6.31Hz": {
        "section": "SAOS parameters",
        "parameter": "Loss tangent at 6.31 Hz",
        "symbol_unit": "tan δ6.31Hz, dimensionless",
        "input_set": "formulation only",
        "order": 9,
    },
}

LAOS_TARGETS = {
    "break_strain_pct": ("Breaking strain", "γbreak, %", 1),
    "log10_break_stress_Pa": ("Breaking stress", "log10 σbreak, Pa", 2),
    "LVR_pct": ("Linear viscoelastic region", "LVR, %", 3),
}

LAOS_TASKS = {
    "strain_from_formulation": ("formulation only", 1),
    "strain_from_viscosity": ("formulation + viscosity", 2),
    "strain_from_saos": ("formulation + SAOS", 3),
    "strain_from_visc_saos": ("formulation + viscosity + SAOS", 4),
}


def latest_benchmark() -> Path:
    candidates = sorted(OUTPUTS.glob("ML_results_xanthan_positive_*"), key=lambda p: p.stat().st_mtime)
    if not candidates:
        raise FileNotFoundError("No ML_results_xanthan_positive_* folders found")
    return candidates[-1]


def read_metrics(benchmark: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    external = pd.read_csv(benchmark / "all_external_model_metrics.csv")
    internal = pd.read_csv(benchmark / "all_internal_group_cv_mean_metrics.csv")
    return external, internal


def metric_string(value: float, metric: str) -> str:
    if pd.isna(value):
        return ""
    if metric == "r2":
        return f"{value:.3f}"
    if abs(value) >= 100:
        return f"{value:.1f}"
    if abs(value) >= 10:
        return f"{value:.2f}"
    if abs(value) >= 1:
        return f"{value:.3f}"
    return f"{value:.4f}"


def annotate_rows(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, r in df.iterrows():
        key = f"{r['task']}|{r['target']}"
        if key in PARAMETER_INFO:
            info = PARAMETER_INFO[key].copy()
        elif r["task"] in LAOS_TASKS and r["target"] in LAOS_TARGETS:
            parameter, symbol_unit, target_order = LAOS_TARGETS[r["target"]]
            input_set, input_order = LAOS_TASKS[r["task"]]
            info = {
                "section": "LAOS parameters",
                "parameter": parameter,
                "symbol_unit": symbol_unit,
                "input_set": input_set,
                "order": 200 + target_order * 10 + input_order,
            }
        else:
            continue
        row = {
            **info,
            "task": r["task"],
            "target": r["target"],
            "feature_set": r.get("feature_set", ""),
            "model": MODEL_LABELS.get(r["model"], r["model"]),
            "r2": r["r2"],
            "mae": r["mae"],
            "rmse": r["rmse"],
            "n": r.get("n", None),
        }
        rows.append(row)
    out = pd.DataFrame(rows)
    out["model"] = pd.Categorical(out["model"], categories=[m for m in MODEL_ORDER if m in set(out["model"])], ordered=True)
    return out.sort_values(["order", "model"]).reset_index(drop=True)


def add_best_flags(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["best_r2"] = False
    df["best_mae"] = False
    df["best_rmse"] = False
    for _, idx in df.groupby(["task", "target", "input_set"], observed=False).groups.items():
        group = df.loc[list(idx)]
        if group["r2"].notna().any():
            df.loc[group["r2"].idxmax(), "best_r2"] = True
        if group["mae"].notna().any():
            df.loc[group["mae"].idxmin(), "best_mae"] = True
        if group["rmse"].notna().any():
            df.loc[group["rmse"].idxmin(), "best_rmse"] = True
    return df


def to_display(df: pd.DataFrame, markdown: bool = False) -> pd.DataFrame:
    rows = []
    for _, r in df.iterrows():
        r2 = metric_string(r["r2"], "r2")
        mae = metric_string(r["mae"], "mae")
        rmse = metric_string(r["rmse"], "rmse")
        if markdown:
            if r["best_r2"]:
                r2 = f"**{r2}**"
            if r["best_mae"]:
                mae = f"**{mae}**"
            if r["best_rmse"]:
                rmse = f"**{rmse}**"
        rows.append(
            {
                "Section": r["section"],
                "Parameter": r["parameter"],
                "Symbol and unit": r["symbol_unit"],
                "Input set": r["input_set"],
                "Model": r["model"],
                "R²": r2,
                "MAE": mae,
                "RMSE": rmse,
            }
        )
    return pd.DataFrame(rows)


def write_markdown(df: pd.DataFrame, path: Path, title: str) -> None:
    def md_escape(value) -> str:
        return str(value).replace("|", "\\|")

    def frame_to_markdown(frame: pd.DataFrame) -> str:
        cols = list(frame.columns)
        out = [
            "| " + " | ".join(md_escape(c) for c in cols) + " |",
            "| " + " | ".join("---" for _ in cols) + " |",
        ]
        for _, row in frame.iterrows():
            out.append("| " + " | ".join(md_escape(row[c]) for c in cols) + " |")
        return "\n".join(out)

    lines = [f"# {title}", "", "Bold indicates the best model for a given parameter/input set: highest R², lowest MAE, or lowest RMSE.", ""]
    display = to_display(df, markdown=True)
    for section in ["SAOS parameters", "Viscosity parameters", "LAOS parameters"]:
        sub = display[display["Section"] == section].drop(columns=["Section"])
        if sub.empty:
            continue
        lines.extend([f"## {section}", "", frame_to_markdown(sub), ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def write_csv(df: pd.DataFrame, path: Path) -> None:
    out = to_display(df, markdown=False)
    out.to_csv(path, index=False)


def style_sheet(ws, df: pd.DataFrame, title: str) -> None:
    thin = Side(style="thin", color="D9D9D9")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    best_fill = PatternFill("solid", fgColor="FFF2CC")
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A2"] = "Bold/yellow cells indicate the best value within each parameter/input set: highest R², lowest MAE, or lowest RMSE."
    ws["A2"].font = Font(italic=True, size=11, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    headers = ["Section", "Parameter", "Symbol and unit", "Input set", "Model", "R²", "MAE", "RMSE"]
    for col, header in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=header)
        c.fill = header_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    display = to_display(df, markdown=False)
    last_section = None
    row_num = 4
    raw_idx = 0
    for _, disp in display.iterrows():
        if disp["Section"] != last_section:
            ws.cell(row=row_num, column=1, value=disp["Section"])
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            c = ws.cell(row=row_num, column=1)
            c.fill = section_fill
            c.font = Font(bold=True, color="1F4E78")
            c.alignment = Alignment(horizontal="left")
            row_num += 1
            last_section = disp["Section"]
        for col, key in enumerate(headers, 1):
            value = disp[key]
            if key == "Section":
                value = ""
            c = ws.cell(row=row_num, column=col, value=value)
            c.border = Border(bottom=thin)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if key in {"R²", "MAE", "RMSE"}:
                c.alignment = Alignment(horizontal="center", vertical="top")
        raw = df.iloc[raw_idx]
        for metric_col, flag in [(6, "best_r2"), (7, "best_mae"), (8, "best_rmse")]:
            if raw[flag]:
                c = ws.cell(row=row_num, column=metric_col)
                c.font = Font(bold=True)
                c.fill = best_fill
        raw_idx += 1
        row_num += 1
    widths = {
        "A": 3,
        "B": 28,
        "C": 34,
        "D": 32,
        "E": 20,
        "F": 10,
        "G": 10,
        "H": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 22


def write_xlsx(external: pd.DataFrame, internal: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "External validation"
    style_sheet(ws, external, "External validation performance")
    ws2 = wb.create_sheet("Internal 5-fold CV")
    style_sheet(ws2, internal, "Internal formulation-based 5-fold CV performance")
    ws3 = wb.create_sheet("Notes")
    notes = [
        ["Item", "Description"],
        ["Validation", "External validation uses held-out prediction formulations. Internal validation uses 5-fold GroupKFold by formulation."],
        ["Bold/yellow cells", "Best model within each parameter/input set for R², MAE, or RMSE."],
        ["Log targets", "MAE and RMSE are in the transformed target units shown in the Symbol and unit column."],
        ["Source", "Generated from latest ML_results_xanthan_positive_* benchmark metrics."],
    ]
    for r, row in enumerate(notes, 1):
        for c, value in enumerate(row, 1):
            ws3.cell(r, c, value)
    ws3["A1"].font = ws3["B1"].font = Font(bold=True, color="FFFFFF")
    ws3["A1"].fill = ws3["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 100
    for row in ws3.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    benchmark = latest_benchmark()
    external_raw, internal_raw = read_metrics(benchmark)
    external = add_best_flags(annotate_rows(external_raw))
    internal = add_best_flags(annotate_rows(internal_raw))
    external.to_csv(OUT_DIR / "publication_metrics_external_raw_with_best_flags.csv", index=False)
    internal.to_csv(OUT_DIR / "publication_metrics_internal_cv_raw_with_best_flags.csv", index=False)
    write_csv(external, OUT_DIR / "publication_table_external.csv")
    write_csv(internal, OUT_DIR / "publication_table_internal_5fold_cv.csv")
    write_markdown(external, OUT_DIR / "publication_table_external.md", "Publication Table: External Validation")
    write_markdown(internal, OUT_DIR / "publication_table_internal_5fold_cv.md", "Publication Table: Internal 5-fold Formulation CV")
    write_xlsx(external, internal, OUT_DIR / "rheology_ml_publication_metric_table.xlsx")
    shutil.copy2(Path(__file__), OUT_DIR / Path(__file__).name)
    summary = {
        "run_id": RUN_ID,
        "benchmark_source": str(benchmark),
        "outputs": [
            "rheology_ml_publication_metric_table.xlsx",
            "publication_table_external.md",
            "publication_table_internal_5fold_cv.md",
            "publication_table_external.csv",
            "publication_table_internal_5fold_cv.csv",
        ],
        "sections": ["SAOS parameters", "Viscosity parameters", "LAOS parameters"],
        "bold_rule": "highest R2, lowest MAE, and lowest RMSE within each parameter/input set",
    }
    (OUT_DIR / "run_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(OUT_DIR)


if __name__ == "__main__":
    main()
